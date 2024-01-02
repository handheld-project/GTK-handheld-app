import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

class CsvFileMaker:
    def __init__(self):
        self.col_size = { 'A' : 35 , 'B' : 15 , 'C' : 15  , 'D' : 45 , 'E' : 35 , 'F' : 15  , 'G' :15  , 'H' : 15 , 'I' :15 , 'J' : 35 }

    def update_excel(self, document , img_path ):
        date = datetime.strptime(document['time'], "%Y-%m-%d_%H-%M-%S")
        formatted_date = date.strftime("%Y-%m-%d")
        file_name = f"{formatted_date}.xlsx"
        moveable = ''
        img = Image(img_path)
        img.width = 320
        img.height = 180

        if ( document['is_movable']):
            moveable = "ตัวอักษรในป่ายเคลื่อนที่ได้"
        else :
            moveable = "ตัวอักษรในป่ายเคลื่อนที่ไม่ได้"

        row_data = [    img_path , document['width'], document['height'], document['type'], moveable ,
                        document['area'], document['price'], document['latitude'], document['longitude'],
                        document['time'] ]


        # Check if the file already exists
        if os.path.isfile(file_name):
            wb = load_workbook(file_name)
            sheet = wb.active

        else:
            # Create a new workbook and sheet
            wb = Workbook()
            sheet = wb.active

            # Write headers
            header = ["ชื่อรูปภาพ", "ความกว้างป้าย ", "ความสูงป่าย", "ชนิดของป้าย", "ลักษณะของป้าย", "พื้นที่ป้าย", "ราคาภาษี", "ละติจูด", "ลองจิจูด", "เวลาที่เก็บข้อมูล"]
            sheet.append(header)

            # Write data rows
        work_row = sheet.max_row + 1
        sheet.row_dimensions[work_row ].height = 160
        sheet.add_image(img, 'L{}'.format(work_row))
        sheet.append(row_data)
        self.set_column_width(sheet)
        self.align_text_middle_by_row(sheet ,work_row )
        wb.save(file_name)
        print("Create complete")

    def set_column_width(self , sheet ) :
        for colName, width in self.col_size.items():
            sheet.column_dimensions[colName].width = width

    def align_text_middle_by_row(self, sheet , row_index ):

        for col_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.alignment = Alignment(vertical='center', horizontal='center')

